"""
Instant Multi-Asset VaR — drop a file, get VaR in seconds.

Run locally:
    pip install -r requirements.txt
    streamlit run app.py

Handles the exact problem this was built for: a huge Excel file that chokes
native Excel formulas. Large .xlsx files are converted once to Parquet
(cached on disk) so every subsequent run/recompute is instant.
"""

import os
# Must be set BEFORE numpy is imported. On memory-constrained cloud containers
# (like Streamlit Community Cloud's free tier), numpy's underlying BLAS math
# library can try to spin up multiple threads for matrix operations (covariance,
# Cholesky decomposition) and crash with a segmentation fault instead of a
# normal Python error. Forcing single-threaded BLAS avoids this entirely —
# slightly slower per operation, but nowhere near enough to matter at this
# problem size, and it won't crash.
os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("OPENBLAS_NUM_THREADS", "1")
os.environ.setdefault("MKL_NUM_THREADS", "1")
os.environ.setdefault("NUMEXPR_NUM_THREADS", "1")
os.environ.setdefault("VECLIB_MAXIMUM_THREADS", "1")

import time
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st

from var_engine import (
    prices_to_returns, normalize_weights,
    parametric_var, historical_var, monte_carlo_var,
    component_var, backtest_var, individual_asset_var,
    MAX_ASSETS_FOR_FULL_COV,
)

st.set_page_config(page_title="Instant Portfolio VaR", layout="wide")

CACHE_DIR = Path(".var_cache")
CACHE_DIR.mkdir(exist_ok=True)


# ---------------------------------------------------------------------------
# File loading with Parquet caching (this is what kills the 246MB problem)
# ---------------------------------------------------------------------------

@st.cache_data(show_spinner=False)
def list_sheet_names(file_bytes: bytes, filename: str) -> list[str]:
    """
    Peeks at an xlsx/xlsm file's sheet names without loading the full data —
    fast even for large files, since it only reads workbook metadata.
    """
    suffix = Path(filename).suffix.lower()
    if suffix not in (".xlsx", ".xlsm"):
        return []
    tmp_path = CACHE_DIR / f"_peek_{filename}"
    tmp_path.write_bytes(file_bytes)
    import openpyxl
    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


@st.cache_data(show_spinner=False)
def load_prices(file_bytes: bytes, filename: str, header_row: int = 1,
                 date_col_letter: str = "A", fill_method: str = "none",
                 sheet_name: str = None) -> pd.DataFrame:
    """
    Loads a wide dates x assets price file.

    header_row: the row number (1-indexed, as you'd see it in Excel) that
        contains the asset codes/tickers. Everything above it is ignored.
    date_col_letter: the Excel column letter that holds the dates (e.g. "A", "B").
    fill_method: "none", "ffill" (carry last known price forward — the
        correct choice for illiquid instruments like bonds that don't trade
        every day), or "drop" (drop any row with missing values).
    sheet_name: which worksheet to read (None = first sheet).

    Caches a Parquet copy so re-runs (or re-computes with different weights)
    skip the expensive Excel parse entirely.
    """
    suffix = Path(filename).suffix.lower()
    cache_key = f"{filename}__s{sheet_name}_h{header_row}_c{date_col_letter}_f{fill_method}"
    cache_path = CACHE_DIR / f"{cache_key}.parquet"

    if cache_path.exists():
        return pd.read_parquet(cache_path)

    tmp_path = CACHE_DIR / f"_raw_{filename}"
    if not tmp_path.exists():
        tmp_path.write_bytes(file_bytes)

    t0 = time.time()
    if suffix in (".xlsx", ".xlsm"):
        # header_row is 1-indexed in Excel; pandas' `header=` is 0-indexed
        df = pd.read_excel(tmp_path, header=header_row - 1, engine="openpyxl",
                            sheet_name=sheet_name if sheet_name else 0)
    elif suffix == ".csv":
        df = pd.read_csv(tmp_path, header=header_row - 1)
    elif suffix == ".parquet":
        df = pd.read_parquet(tmp_path)
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    # Resolve the date column: pandas gives generic column names for the
    # header row we picked, so find the one matching the chosen letter.
    if suffix != ".parquet":
        col_idx = _excel_col_to_index(date_col_letter)
        date_col_name = df.columns[col_idx]
        df = df.rename(columns={date_col_name: "__date__"}).set_index("__date__")
        # drop any other columns to the left of the date column that were just headers/blank
        df = df.drop(columns=[c for c in df.columns if str(c).startswith("Unnamed") and df[c].isna().all()])

    # dayfirst=True because Indian financial files use dd-mm-yyyy, not mm-dd-yyyy
    df.index = pd.to_datetime(df.index, errors="coerce", dayfirst=True)
    df = df[df.index.notna()]
    df = df.sort_index()

    # "NA" text, blanks, dashes etc. all become proper missing values
    df = df.replace(["NA", "N/A", "-", "", "#N/A"], np.nan)
    df = df.apply(pd.to_numeric, errors="coerce")
    # a price of 0 (or negative) is always a data artifact, never a real bond/stock
    # price — treat it as missing rather than a genuine 100% crash
    df = df.mask(df <= 0)

    # drop columns that are entirely empty (no trades ever recorded)
    df = df.dropna(axis=1, how="all")

    if fill_method == "ffill":
        df = df.ffill()
    elif fill_method == "drop":
        df = df.dropna(axis=0, how="any")

    df.to_parquet(cache_path)

    st.session_state["_last_load_seconds"] = round(time.time() - t0, 1)
    return df


@st.cache_data(show_spinner=False)
def load_prices_long(file_bytes: bytes, filename: str, sheet_name: str,
                      date_col_letter: str, id_col_letter: str, price_col_letter: str,
                      header_row: int = 1, fill_method: str = "none") -> pd.DataFrame:
    """
    Loads a 'long format' file: one row per (date, asset) pair, e.g. a daily
    holdings/valuation report where each security reappears on every date it
    was held. Pivots into the same wide dates x assets shape the rest of the
    app expects.

    Uses openpyxl in streaming read-only mode (iter_rows) rather than
    pandas.read_excel. This matters a lot on large files: pandas' Excel
    reader builds the full workbook object in memory regardless of usecols,
    which can need 5-10x the file size in RAM — enough to crash a
    memory-limited server on a 200MB+ file. Streaming mode reads one row at
    a time and only keeps the 3 columns we actually need.
    """
    import openpyxl

    tmp_path = CACHE_DIR / f"_raw_{filename}"
    if not tmp_path.exists():
        tmp_path.write_bytes(file_bytes)

    cache_key = f"{filename}__long_{sheet_name}_{date_col_letter}_{id_col_letter}_{price_col_letter}_f{fill_method}"
    cache_path = CACHE_DIR / f"{cache_key}.parquet"
    if cache_path.exists():
        return pd.read_parquet(cache_path)

    t0 = time.time()

    date_idx = _excel_col_to_index(date_col_letter)
    id_idx = _excel_col_to_index(id_col_letter)
    price_idx = _excel_col_to_index(price_col_letter)
    max_idx = max(date_idx, id_idx, price_idx)

    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    dates, asset_ids, prices = [], [], []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) <= max_idx:
            continue
        d = row[date_idx]
        if d is None:
            continue
        dates.append(d)
        asset_ids.append(row[id_idx])
        prices.append(row[price_idx])
    wb.close()

    df = pd.DataFrame({"date": dates, "asset_id": asset_ids, "price": prices})

    df["date"] = pd.to_datetime(df["date"], errors="coerce", dayfirst=True)
    df = df[df["date"].notna()]
    df["asset_id"] = df["asset_id"].astype(str).str.strip()
    df["price"] = df["price"].replace(["NA", "N/A", "-", "", "#N/A"], np.nan)
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    # a price of 0 (or negative) is always a data artifact — treat as missing
    df.loc[df["price"] <= 0, "price"] = np.nan

    # if a date+asset pair appears more than once, keep the last recorded value
    df = df.drop_duplicates(subset=["date", "asset_id"], keep="last")

    wide = df.pivot(index="date", columns="asset_id", values="price")
    wide = wide.sort_index()
    wide = wide.dropna(axis=1, how="all")

    if fill_method == "ffill":
        wide = wide.ffill()
    elif fill_method == "drop":
        wide = wide.dropna(axis=0, how="any")

    wide.to_parquet(cache_path)
    st.session_state["_last_load_seconds"] = round(time.time() - t0, 1)
    return wide


@st.cache_data(show_spinner=False)
def load_long_asof(file_bytes: bytes, filename: str, sheet_name: str,
                    date_col_letter: str, id_col_letter: str, price_col_letter: str,
                    header_row: int, as_of_date: "pd.Timestamp", lookback_years: int,
                    fill_method: str = "ffill") -> tuple:
    """
    Like load_prices_long, but restricted to:
      1. Only securities that have an actual row on exactly `as_of_date`
         (i.e. still held on that date — matured/sold securities are excluded
         entirely, even from the lookback window).
      2. Only the `lookback_years` of history immediately before as_of_date
         (out-of-window rows are skipped during streaming, not just filtered
         afterward, to keep memory low on a large file).

    Returns (wide_price_df, held_isins_count).
    """
    import openpyxl
    import datetime as dt

    tmp_path = CACHE_DIR / f"_raw_{filename}"
    if not tmp_path.exists():
        tmp_path.write_bytes(file_bytes)

    cache_key = (f"{filename}__asof_{sheet_name}_{date_col_letter}_{id_col_letter}_"
                 f"{price_col_letter}_{as_of_date.date()}_{lookback_years}y_f{fill_method}")
    cache_path = CACHE_DIR / f"{cache_key}.parquet"
    meta_path = CACHE_DIR / f"{cache_key}.heldcount"
    if cache_path.exists() and meta_path.exists():
        return pd.read_parquet(cache_path), int(meta_path.read_text())

    t0 = time.time()

    date_idx = _excel_col_to_index(date_col_letter)
    id_idx = _excel_col_to_index(id_col_letter)
    price_idx = _excel_col_to_index(price_col_letter)
    max_idx = max(date_idx, id_idx, price_idx)

    as_of_norm = pd.Timestamp(as_of_date).normalize()
    window_start = as_of_norm - pd.DateOffset(years=lookback_years)

    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    dates, asset_ids, prices = [], [], []
    held_isins = set()

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) <= max_idx:
            continue
        d_raw = row[date_idx]
        if d_raw is None:
            continue
        if isinstance(d_raw, (dt.datetime, dt.date)):
            d = pd.Timestamp(d_raw).normalize()
        else:
            d = pd.to_datetime(d_raw, errors="coerce", dayfirst=True)
            if pd.isna(d):
                continue
            d = d.normalize()

        if d == as_of_norm:
            held_isins.add(str(row[id_idx]).strip())

        if window_start <= d <= as_of_norm:
            dates.append(d)
            asset_ids.append(str(row[id_idx]).strip())
            prices.append(row[price_idx])
    wb.close()

    df = pd.DataFrame({"date": dates, "asset_id": asset_ids, "price": prices})
    # keep only securities actually held on the as-of date
    df = df[df["asset_id"].isin(held_isins)]

    df["price"] = df["price"].replace(["NA", "N/A", "-", "", "#N/A"], np.nan)
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df.loc[df["price"] <= 0, "price"] = np.nan
    df = df.drop_duplicates(subset=["date", "asset_id"], keep="last")

    wide = df.pivot(index="date", columns="asset_id", values="price")
    wide = wide.sort_index()
    wide = wide.dropna(axis=1, how="all")

    if fill_method == "ffill":
        wide = wide.ffill()
    elif fill_method == "drop":
        wide = wide.dropna(axis=0, how="any")

    wide.to_parquet(cache_path)
    meta_path.write_text(str(len(held_isins)))
    st.session_state["_last_load_seconds"] = round(time.time() - t0, 1)
    return wide, len(held_isins)


@st.cache_data(show_spinner=False)
def preview_columns(file_bytes: bytes, filename: str, sheet_name: str, header_row: int,
                     date_col_letter: str, id_col_letter: str, price_col_letter: str) -> pd.DataFrame:
    """
    Reads just the header + first 5 data rows for the 3 chosen columns, so
    you can visually confirm the letters point to the right data BEFORE
    running the full computation on a large file. Uses streaming read-only
    mode and stops after 5 rows — never touches the rest of the file.
    """
    import openpyxl

    tmp_path = CACHE_DIR / f"_raw_{filename}"
    if not tmp_path.exists():
        tmp_path.write_bytes(file_bytes)

    date_idx = _excel_col_to_index(date_col_letter)
    id_idx = _excel_col_to_index(id_col_letter)
    price_idx = _excel_col_to_index(price_col_letter)
    max_idx = max(date_idx, id_idx, price_idx)

    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True)):
        if len(row) <= max_idx:
            continue
        rows.append((row[date_idx], row[id_idx], row[price_idx]))
        if len(rows) >= 5:
            break
    wb.close()

    return pd.DataFrame(rows, columns=["date", "asset_id (ISIN etc.)", "price"])


def _excel_col_to_index(letter: str) -> int:
    """Converts an Excel column letter ('A', 'B', ..., 'AA', ...) to a 0-indexed position."""
    letter = letter.strip().upper()
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


def make_synthetic_demo(n_days=1000, n_assets=8, seed=7) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    tickers = [f"STOCK_{i+1}" for i in range(n_assets)]
    market = rng.normal(0.0005, 0.01, n_days)
    betas = rng.uniform(0.4, 1.6, n_assets)
    idio = rng.normal(0, 0.015, (n_days, n_assets))
    daily_returns = market[:, None] * betas[None, :] + idio
    prices = 100 * np.cumprod(1 + daily_returns, axis=0)
    return pd.DataFrame(prices, columns=tickers,
                         index=pd.date_range("2022-01-01", periods=n_days, freq="B"))


# ---------------------------------------------------------------------------
# Sidebar — inputs
# ---------------------------------------------------------------------------

st.title("📉 Instant Multi-Asset Portfolio VaR")
st.caption("Drop a price file → get Parametric, Historical & Monte Carlo VaR in seconds — no Excel recalculation.")

with st.sidebar:
    st.header("1. Data")
    use_demo = st.checkbox("Use synthetic demo data instead of uploading", value=True)

    uploaded = None
    header_row, date_col_letter, fill_method, sheet_name = 1, "A", "none", None
    data_format = "wide"
    id_col_letter, price_col_letter = "C", "V"
    asof_mode, as_of_date, lookback_years = False, None, 3
    if not use_demo:
        uploaded = st.file_uploader(
            "Upload price file",
            type=["xlsx", "xlsm", "csv", "parquet"],
        )
        st.caption(
            "Large .xlsx/.xlsm files are cached to Parquet after the first load — "
            "every recompute after that is instant."
        )

        sheet_name = None
        if uploaded is not None and Path(uploaded.name).suffix.lower() in (".xlsx", ".xlsm"):
            sheet_names = list_sheet_names(uploaded.getvalue(), uploaded.name)
            if len(sheet_names) > 1:
                sheet_name = st.selectbox(
                    "Which worksheet has the actual price data?",
                    options=sheet_names,
                )
            elif sheet_names:
                sheet_name = sheet_names[0]

        data_format = st.radio(
            "How is your data laid out?",
            options=["wide", "long"],
            format_func=lambda x: {
                "wide": "Wide — one row per date, one column per asset",
                "long": "Long — one row per date+security (e.g. a holdings/valuation report)",
            }[x],
            index=0,
        )

        if data_format == "long":
            st.caption("Only these 3 columns will be read — much faster on a large file.")
            date_col_letter = st.text_input("Date column letter", value="A")
            id_col_letter = st.text_input("Security ID column letter (e.g. ISIN)", value="C")
            price_col_letter = st.text_input("Price column letter (e.g. Market Rate)", value="V")
            header_row = st.number_input("Header row", min_value=1, value=1, step=1)
            fill_method = st.selectbox(
                "How to handle missing/'NA' values?",
                options=["none", "ffill", "drop"],
                format_func=lambda x: {
                    "none": "Leave as missing (default)",
                    "ffill": "Carry forward last known price — use this for bonds/illiquid assets that don't trade daily",
                    "drop": "Drop any date with a missing value in any asset",
                }[x],
                index=1,
            )

            asof_mode = st.checkbox(
                "Restrict to securities currently held as of a specific date (recommended for 'current portfolio' analysis)",
                value=False,
            )
            as_of_date, lookback_years = None, 3
            if asof_mode:
                as_of_date = st.date_input("As-of date", value=pd.Timestamp("2026-03-31"))
                lookback_years = st.number_input("Years of history to use", min_value=1, max_value=10, value=3, step=1)
                st.caption(
                    "Only securities with an actual row on this exact date count as 'held' — "
                    "matured/sold securities are excluded, even from the lookback window."
                )

            if uploaded is not None and sheet_name:
                try:
                    preview_df = preview_columns(
                        uploaded.getvalue(), uploaded.name, sheet_name, header_row,
                        date_col_letter, id_col_letter, price_col_letter,
                    )
                    st.caption("Preview — check this looks right before computing:")
                    st.dataframe(preview_df, hide_index=True, width='stretch')
                except Exception as e:
                    st.error(f"Couldn't preview with these column letters: {e}")
        else:
            with st.expander("File layout settings (open this if your file has extra header rows, a different date column, or gaps like 'NA')"):
                header_row = st.number_input(
                    "Which row has the asset names/codes (as seen in Excel)?",
                    min_value=1, value=1, step=1,
                )
                date_col_letter = st.text_input(
                    "Which column has the dates (Excel letter, e.g. A, B, C)?",
                    value="A",
                )
                fill_method = st.selectbox(
                    "How to handle missing/'NA' values?",
                    options=["none", "ffill", "drop"],
                    format_func=lambda x: {
                        "none": "Leave as missing (default)",
                        "ffill": "Carry forward last known price — use this for bonds/illiquid assets that don't trade daily",
                        "drop": "Drop any date with a missing value in any asset",
                    }[x],
                    index=0,
                )

    st.header("2. VaR settings")
    confidence = st.select_slider("Confidence level", options=[0.90, 0.95, 0.975, 0.99], value=0.95)
    horizon_days = st.number_input("Horizon (days)", min_value=1, max_value=30, value=1)
    portfolio_value = st.number_input("Portfolio value (₹)", min_value=0.0, value=10_000_000.0, step=100000.0, format="%.0f")
    n_sims = st.number_input("Monte Carlo simulations", min_value=1000, max_value=100000, value=5000, step=1000)
    st.caption("Lower this if you have a very large number of assets and hit memory errors on the deployed app.")


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------

if use_demo:
    prices = make_synthetic_demo()
    st.info("Using synthetic demo data (8 correlated assets, 1000 trading days). Uncheck the box in the sidebar to upload your own file.")
elif uploaded is not None:
    if data_format == "long":
        if not sheet_name:
            st.error("Could not determine the sheet name — please check the file.")
            st.stop()
        if asof_mode and as_of_date:
            prices, held_count = load_long_asof(
                uploaded.getvalue(), uploaded.name, sheet_name,
                date_col_letter, id_col_letter, price_col_letter,
                header_row, pd.Timestamp(as_of_date), lookback_years, fill_method,
            )
            st.info(f"{held_count} securities were held on {pd.Timestamp(as_of_date).date()} and are included below "
                    f"(using {lookback_years} year(s) of history).")
        else:
            prices = load_prices_long(
                uploaded.getvalue(), uploaded.name, sheet_name,
                date_col_letter, id_col_letter, price_col_letter,
                header_row, fill_method,
            )
    else:
        prices = load_prices(uploaded.getvalue(), uploaded.name, header_row, date_col_letter, fill_method, sheet_name)
    load_time = st.session_state.get("_last_load_seconds")
    if load_time:
        st.success(f"Loaded and cached to Parquet in {load_time}s. Future recomputes on this file will be instant.")
    if prices.shape[1] == 0:
        st.error("No asset columns were found. Double-check the column letter / header row settings above.")
        st.stop()
else:
    st.warning("Upload a file or check 'use synthetic demo data' to continue.")
    st.stop()

returns, dropped_assets = prices_to_returns(prices)
tickers = list(returns.columns)

st.subheader("Loaded data")
c1, c2, c3 = st.columns(3)
c1.metric("Assets", len(tickers))
c2.metric("Trading days", len(returns))
c3.metric("Date range", f"{returns.index.min().date()} → {returns.index.max().date()}")

if len(tickers) > MAX_ASSETS_FOR_FULL_COV:
    st.error(
        f"{len(tickers)} securities are loaded, which is more than this server's "
        f"memory can safely handle for portfolio VaR ({MAX_ASSETS_FOR_FULL_COV} max) — "
        f"clicking Compute would likely crash the app. Restrict to a smaller set of "
        f"securities (e.g. top positions by size) or run this app locally/on a "
        f"larger machine, where this limit doesn't apply."
    )

if dropped_assets:
    st.warning(
        f"{len(dropped_assets)} asset(s) were excluded — not enough price history "
        f"(need at least 30 valid price points) to reliably estimate risk. "
        f"Excluded: {', '.join(dropped_assets[:15])}"
        + (f", and {len(dropped_assets)-15} more" if len(dropped_assets) > 15 else "")
    )

with st.expander("Preview returns matrix"):
    st.dataframe(returns.tail(10), width='stretch')


# ---------------------------------------------------------------------------
# Individual Security VaR (only shown in as-of-date mode)
# ---------------------------------------------------------------------------

if asof_mode:
    st.subheader("2. Individual Security VaR — Top 5 riskiest holdings")
    st.caption(
        "Standalone VaR per security (its own volatility, no portfolio diversification "
        "benefit factored in) — ranks currently-held securities by risk on their own."
    )
    ind_var = individual_asset_var(returns, confidence)
    top5 = ind_var.head(5).copy()
    top5.index.name = "Security ID"
    st.dataframe(
        top5.style.format({
            "historical_var_pct": "{:.3%}",
            "parametric_var_pct": "{:.3%}",
        }),
        width='stretch',
    )
    with st.expander(f"See all {len(ind_var)} held securities ranked"):
        st.dataframe(
            ind_var.style.format({"historical_var_pct": "{:.3%}", "parametric_var_pct": "{:.3%}"}),
            width='stretch',
        )


# ---------------------------------------------------------------------------
# Weights
# ---------------------------------------------------------------------------

st.subheader("3. Portfolio weights")
weight_mode = st.radio("Set weights", ["Equal weight", "Manual entry"], horizontal=True)

if weight_mode == "Equal weight":
    raw_weights = {t: 1.0 for t in tickers}
else:
    default_df = pd.DataFrame({"ticker": tickers, "weight": [1.0] * len(tickers)})
    edited = st.data_editor(default_df, hide_index=True, width='stretch', key="weights_editor")
    raw_weights = dict(zip(edited["ticker"], edited["weight"]))

weights = normalize_weights(raw_weights)

if not st.button("🚀 Compute VaR", type="primary"):
    st.stop()


# ---------------------------------------------------------------------------
# Compute — this whole block should run in well under a second
# ---------------------------------------------------------------------------

t0 = time.time()

try:
    p_res, cov_matrix, marginal_var = parametric_var(returns, weights, portfolio_value, confidence, horizon_days)
    h_res = historical_var(returns, weights, portfolio_value, confidence, horizon_days)
    mc_res = monte_carlo_var(returns, weights, portfolio_value, confidence, horizon_days, n_sims=n_sims)

    w_aligned = weights.reindex(returns.columns).fillna(0.0).astype(np.float32)
    port_vol_daily = float(np.sqrt(w_aligned.values @ cov_matrix.values @ w_aligned.values.T))
    comp = component_var(marginal_var, weights, portfolio_value, p_res.var_pct, port_vol_daily)

    bt = backtest_var(returns, weights, p_res.var_pct)
except MemoryError as e:
    # This is exactly the failure mode that used to crash the whole server
    # with an unrecoverable segmentation fault (too many held assets for a
    # dense covariance matrix to fit in memory). var_engine now raises this
    # as a normal exception *before* attempting the dangerous allocation, so
    # we can show a clear on-screen message instead of the app dying.
    st.error(str(e))
    st.stop()

compute_time = time.time() - t0

st.success(f"Computed all three VaR methods + component breakdown + backtest in {compute_time:.3f} seconds.")

st.subheader(f"4. Results — {int(confidence*100)}% confidence, {horizon_days}-day horizon")

col1, col2, col3 = st.columns(3)
for col, res in zip([col1, col2, col3], [p_res, h_res, mc_res]):
    col.metric(res.method, f"{res.var_pct*100:.2f}%", f"₹ {res.var_amount:,.0f}")

st.caption(
    "If the three methods diverge a lot, your return distribution likely has fat tails "
    "or skew that the parametric (normal-distribution) method underestimates — trust "
    "Historical/Monte Carlo more in that case."
)

st.subheader("5. Component VaR — which holdings drive the risk")
st.dataframe(
    comp.style.format({
        "weight": "{:.2%}",
        "component_var_pct": "{:.3%}",
        "component_var_amount": "₹ {:,.0f}",
        "pct_of_total_var": "{:.1%}",
    }),
    width='stretch',
)
st.bar_chart(comp["component_var_amount"])

st.subheader("6. Backtest — did actual losses exceed VaR as often as expected?")
expected_breach_rate = 1 - confidence
bcol1, bcol2, bcol3 = st.columns(3)
bcol1.metric("Trading days tested", bt["total_days"])
bcol2.metric("Breaches", bt["breaches"])
bcol3.metric("Actual breach rate", f"{bt['breach_rate']*100:.2f}%", f"expected ≈ {expected_breach_rate*100:.1f}%")

if bt["breach_rate"] > expected_breach_rate * 1.5:
    st.warning("Breach rate is notably higher than expected — the VaR model may be understating risk.")
elif bt["breach_rate"] < expected_breach_rate * 0.5:
    st.info("Breach rate is notably lower than expected — the VaR model may be overly conservative.")
